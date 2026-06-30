import openpyxl
from typing import List, Dict
from pathlib import Path


class ExcelParser:
    """Парсер Excel-файлов"""

    @staticmethod
    def parse_students(file_path: str) -> List[Dict]:
        """Парсит файл со студентами"""
        if not Path(file_path).exists():
            raise FileNotFoundError(f"Файл {file_path} не найден")
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        students = []
        # Предполагаем: A — ФИО, B — Курс, C — Группа
        for row in ws.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовки
            if row[0] is None:
                continue
            students.append({
                "full_name": str(row[0]).strip(),
                "course": int(row[1]) if row[1] else 3,
                "group_name": str(row[2]).strip() if row[2] else ""
            })
        return students

    @staticmethod
    def parse_teachers(file_path: str) -> List[Dict]:
        """Парсит файл с преподавателями"""
        if not Path(file_path).exists():
            raise FileNotFoundError(f"Файл {file_path} не найден")
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        teachers = []
        # Предполагаем: A — ФИО, B — Должность, C — Степень, D — Контакт
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            teachers.append({
                "full_name": str(row[0]).strip(),
                "position": str(row[1]).strip() if row[1] else "",
                "degree": str(row[2]).strip() if row[2] else "",
                "contact": str(row[3]).strip() if row[3] else ""
            })
        return teachers

    @staticmethod
    def export_to_excel(data: List[Dict], headers: List[str], file_path: str):
        """Экспортирует данные в Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Заголовки
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Данные
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))
        
        wb.save(file_path)
        return file_path