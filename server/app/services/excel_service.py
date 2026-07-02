import openpyxl
from typing import List, Dict
import os


class ExcelService:
    """Сервис для работы с Excel-файлами"""

    @staticmethod
    def parse_students(file_path: str) -> List[Dict]:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Файл {file_path} не найден")

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        students = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            students.append({
                "full_name": str(row[0]).strip(),
                "course": int(row[1]) if row[1] else 3,
                "group_name": str(row[2]).strip() if row[2] else ""
            })
        return students

    @staticmethod
    def parse_teachers(file_path: str) -> List[Dict]:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Файл {file_path} не найден")

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        teachers = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            teachers.append({
                "full_name": str(row[0]).strip(),
                "position": str(row[1]).strip() if row[1] else "",
                "degree": str(row[2]).strip() if row[2] else "",
                "contact": str(row[3]).strip() if row[3] else ""
            })
        return teachers