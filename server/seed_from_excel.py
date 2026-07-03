# seed_from_excel.py
import sys
from pathlib import Path
sys.path.append(str(Path(__file__).parent))

import openpyxl
from app.database.db import SessionLocal
from app.models import models
from app.core.security import hash_password


def is_name(text):
    """Проверяет, похоже ли текст на ФИО"""
    if not text or not isinstance(text, str):
        return False
    text = text.strip()
    words = text.split()
    if len(words) < 2:
        return False
    # Проверяем, что все слова начинаются с заглавной буквы
    for word in words:
        if not word[0].isupper():
            return False
    # Исключаем мусор
    if text in ["ФИО", "№ п/п", "Группа", "Результаты тестирования", "УПРАВЛЕНИЕ ПЕРСОНАЛОМ", "РЕКЛАМА", "ТУРИЗМ", "СЕРВИС"]:
        return False
    return True


def load_students_from_excel(file_path: str):
    db = SessionLocal()
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        added = 0
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n📄 Лист: {sheet_name}")
            
            # Определяем группу из названия листа
            group = sheet_name
            if "РПО" in sheet_name:
                group = sheet_name.replace("РПО", "").strip()
            elif "Дизайн" in sheet_name:
                group = sheet_name.replace("Дизайн", "").strip()
            elif sheet_name in ["УП", "РиСО", "Туризм", "Сервис"]:
                # Для этих листов группа не нужна, это другие специальности
                # Пропускаем их или назначаем группу по названию
                group = sheet_name
            elif "ПИ" in sheet_name:
                group = "ПИ"
            
            # На листе ПИ данные начинаются со строки 4
            start_row = 4 if sheet_name == "ПИ" else 1
            
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                if not row:
                    continue
                
                # Ищем имя в колонке B (индекс 1) или в первой колонке, если там имя
                name = None
                for idx, cell in enumerate(row):
                    if cell and is_name(str(cell)):
                        name = str(cell).strip()
                        break
                
                if not name:
                    continue
                
                # Если имя начинается с цифры (например, "1. Агеенкова"), убираем цифру
                if name[0].isdigit():
                    parts = name.split(maxsplit=1)
                    if len(parts) > 1:
                        name = parts[1]
                
                # Если это не студент (например, "ПИ в дизайне"), пропускаем
                if name in ["ПИ в дизайне", "ПИ РПО"]:
                    continue
                
                login = name.split()[0].lower()
                
                existing = db.query(models.User).filter(
                    models.User.login == login
                ).first()
                
                if existing:
                    print(f"⏭️ Пропущен: {name} (уже существует)")
                    continue
                
                user = models.User(
                    full_name=name,
                    login=login,
                    hashed_password=hash_password("password"),
                    role="student"
                )
                db.add(user)
                db.flush()
                
                if hasattr(models, "Student"):
                    student = models.Student(
                        id=user.id,
                        course=1,
                        group_name=group
                    )
                    db.add(student)
                
                added += 1
                print(f"✅ Добавлен: {name} (группа: {group})")
        
        db.commit()
        print(f"\n🎉 ВСЕГО ДОБАВЛЕНО: {added} студентов")
        
    except Exception as e:
        db.rollback()
        print(f"❌ ОШИБКА: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()


if __name__ == "__main__":
    file_path = "uploads/Ochnoe_1_kurs_zachislenye_1.xlsx"
    load_students_from_excel(file_path)