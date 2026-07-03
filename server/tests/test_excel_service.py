# tests/test_excel_service.py
import pytest
import openpyxl
import os
from app.services.excel_service import ExcelService


def test_parse_students():
    """Тест: парсинг студентов из Excel"""
    # Создаём тестовый файл
    file_path = "test_students.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "ФИО"
    ws["B1"] = "Курс"
    ws["C1"] = "Группа"
    ws["A2"] = "Бавлов Сергей Александрович"
    ws["B2"] = 1
    ws["C2"] = "14121"
    ws["A3"] = "Ганжитова Ирина Алдаровна"
    ws["B3"] = 1
    ws["C3"] = "14124"
    wb.save(file_path)
    
    students = ExcelService.parse_students(file_path)
    assert len(students) == 2
    assert students[0]["full_name"] == "Бавлов Сергей Александрович"
    assert students[0]["course"] == 1
    assert students[0]["group_name"] == "14121"
    
    os.remove(file_path)


def test_parse_teachers():
    """Тест: парсинг преподавателей из Excel"""
    file_path = "test_teachers.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "ФИО"
    ws["B1"] = "Должность"
    ws["C1"] = "Степень"
    ws["D1"] = "Контакт"
    ws["A2"] = "Иванов Иван Иванович"
    ws["B2"] = "Доцент"
    ws["C2"] = "к.т.н."
    ws["D2"] = "ivanov@email.com"
    wb.save(file_path)
    
    teachers = ExcelService.parse_teachers(file_path)
    assert len(teachers) == 1
    assert teachers[0]["full_name"] == "Иванов Иван Иванович"
    assert teachers[0]["position"] == "Доцент"
    assert teachers[0]["degree"] == "к.т.н."
    assert teachers[0]["contact"] == "ivanov@email.com"
    
    os.remove(file_path)


def test_parse_students_file_not_found():
    """Тест: ошибка при отсутствии файла"""
    with pytest.raises(FileNotFoundError):
        ExcelService.parse_students("non_existent.xlsx")
